"""Erzeugt realistische Beispielangebote zum Ausprobieren.

    python sample_data/erzeuge_beispiele.py

Die Dateien landen in ``sample_data/erzeugt/``.  Sie decken bewusst sehr
unterschiedliche Formate ab -- genau das ist im Alltag das Problem:

    1. Klassische Excel-Preisliste mit Kopfzeile
    2. Excel OHNE Kopfzeile (Spalten nur ueber Datentypen erkennbar)
    3. PDF-Angebot mit Spaltenlayout
    4. E-Mail (.eml) mit formloser Preismitteilung im Text
    5. E-Mail (.eml) mit Excel-Anhang
    6. Reiner Text (fuer "Text einfuegen")
    7. CSV mit Semikolon und deutschen Zahlen
    8. Excel mit Staffelpreisen und Summenzeile (Stoerfaelle)

Dazu kommen die *schwierigen* Faelle -- die, an denen ein Import im Alltag
tatsaechlich scheitert:

    9.  PDF im Fliesstextstil: keine Tabelle, die Preise stehen in Saetzen,
        die Kopfdaten sind ueber das Anschreiben verstreut
    10. Excel mit quer verteilten Kopfdaten: Angebotsnummer in H2, Datum in
        B15 *unter* der Tabelle, Waehrung nur in einer Fusszeile, Kopfzeile
        erst in Zeile 12 -- davor Logo-Platzhalter und Anschreiben
    11. E-Mail mit Anhang, bei der die Mail die entscheidenden Ergaenzungen
        traegt (Gueltigkeit, Streichung, Zusatzposition, ueberholter Preis)
    12. PDF mit Staffelpreisen ueber zwei Seiten inkl. Uebertragszeile
    13. Excel mit vertauschten Artikelnummernspalten, englischen
        Ueberschriften und gemischten Zahlenformaten (1,234.56 und 1.234,56)

Quer durch die schwierigen Faelle laufen zusaetzlich: Umlaute und
Sonderzeichen in den Bezeichnungen, Mengeneinheiten, die erst normalisiert
werden muessen (Stk., St, PCS, Meter), und eine Position "auf Anfrage" ohne
Preis -- dort darf nie ein Betrag erfunden werden.

Die verwendeten Materialnummern passen zum eingebauten Testsystem
(Mock-SAP), damit der Alt/Neu-Vergleich sofort etwas anzeigt.
"""

from __future__ import annotations

import csv
import sys
from datetime import date, timedelta
from email.message import EmailMessage
from pathlib import Path

ZIEL = Path(__file__).resolve().parent / "erzeugt"

HEUTE = date.today()
GUELTIG_AB = (HEUTE.replace(day=1) + timedelta(days=32)).replace(day=1)


def _datum(tag: date) -> str:
    return tag.strftime("%d.%m.%Y")


# ---------------------------------------------------------------------------
# 1 + 2 + 8: Excel
# ---------------------------------------------------------------------------

def excel_mit_kopfzeile(pfad: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = "Preisliste 2026"

    blatt["A1"] = "Muster Dichtungstechnik GmbH"
    blatt["A1"].font = Font(bold=True, size=14)
    blatt["A2"] = "Industriestrasse 14, 33602 Bielefeld"
    blatt["A4"] = "Angebot Nr.:"
    blatt["B4"] = "ANG-2026-04711"
    blatt["A5"] = "Angebotsdatum:"
    blatt["B5"] = _datum(HEUTE)
    blatt["A6"] = "Gueltig bis:"
    blatt["B6"] = _datum(HEUTE + timedelta(days=90))
    blatt["A7"] = "Waehrung:"
    blatt["B7"] = "EUR"
    blatt["D4"] = "Zahlungsbedingungen:"
    blatt["E4"] = "30 Tage netto, 2 % Skonto bei 10 Tagen"
    blatt["D5"] = "Incoterm:"
    blatt["E5"] = "FCA Bielefeld"
    blatt["D6"] = "Ansprechpartner:"
    blatt["E6"] = "T. Wagner, vertrieb@muster-dichtungstechnik.de"

    kopf = ["Pos.", "Materialnummer", "Artikelnummer Lieferant", "Bezeichnung",
            "Menge", "ME", "Preis", "PE", "Mindestmenge", "Lieferzeit",
            "Gueltig ab", "Bemerkung"]
    for spalte, text in enumerate(kopf, start=1):
        zelle = blatt.cell(row=9, column=spalte, value=text)
        zelle.font = Font(bold=True)
        zelle.alignment = Alignment(horizontal="center")

    zeilen = [
        (10, "47110001", "DR-40527-NBR", "Dichtring NBR 40x52x7", 500, "St", "12,85", 1,
         50, "14 Tage", _datum(GUELTIG_AB), "Preis +3,6 %"),
        (20, "47110002", "OR-2503-FPM", "O-Ring Viton 25x3", 2000, "St", "8,90", 10,
         500, "14 Tage", _datum(GUELTIG_AB), "Preis je 10 Stueck"),
        (30, "47110003", "WDR-30477", "Wellendichtring FPM 30x47x7", 250, "St", "18,95", 1,
         25, "21 Tage", _datum(GUELTIG_AB), "unveraendert"),
        (40, "47110004", "FD-10060-2", "Flachdichtung Klingersil 100x60x2", 800, "St",
         "3,40", 1, 100, "10 Tage", _datum(GUELTIG_AB), "NEU im Sortiment"),
        (50, "", "SP-99-001", "Sonderdichtung nach Zeichnung", 20, "St", "145,00", 1,
         5, "35 Tage", _datum(GUELTIG_AB), "Materialnummer bitte ergaenzen"),
    ]
    for index, zeile in enumerate(zeilen, start=10):
        for spalte, wert in enumerate(zeile, start=1):
            blatt.cell(row=index, column=spalte, value=wert)

    for spalte, breite in zip("ABCDEFGHIJKL",
                              (6, 16, 22, 34, 8, 6, 10, 6, 12, 12, 12, 26)):
        blatt.column_dimensions[spalte].width = breite

    mappe.save(pfad)


def excel_ohne_kopfzeile(pfad: Path) -> None:
    """Preisliste ohne jede Ueberschrift -- nur ueber Datentypen erkennbar."""
    from openpyxl import Workbook

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = "Preise"

    zeilen = [
        ("47110005", "Kugellager 6204-2RS", 1000, "ST", "4,55", 1, _datum(GUELTIG_AB)),
        ("49900010", "Hydraulikschlauch DN12 2SN", 300, "M", "7,20", 1, _datum(GUELTIG_AB)),
        ("48200111", "Gleitringdichtung KP-40", 40, "ST", "289,00", 1, _datum(GUELTIG_AB)),
    ]
    for index, zeile in enumerate(zeilen, start=1):
        for spalte, wert in enumerate(zeile, start=1):
            blatt.cell(row=index, column=spalte, value=wert)
    mappe.save(pfad)


def excel_mit_stoerfaellen(pfad: Path) -> None:
    """Staffelpreise, Zwischensummen, Leerzeilen, englische Zahlen."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = "Quotation"

    blatt["A1"] = "Nordtec Industriebedarf AG"
    blatt["A1"].font = Font(bold=True)
    blatt["A2"] = "Quotation No. Q-2026-8842"
    blatt["A3"] = f"Date: {HEUTE.strftime('%Y-%m-%d')}"
    blatt["A4"] = "Currency: EUR"

    kopf = ["Item", "Material", "Description", "Qty", "Unit", "Unit price",
            "Price unit", "Valid from"]
    for spalte, text in enumerate(kopf, start=1):
        blatt.cell(row=6, column=spalte, value=text).font = Font(bold=True)

    zeilen = [
        (10, "47110005", "Ball bearing 6204-2RS", 100, "PCS", "4.75", 1,
         GUELTIG_AB.strftime("%Y-%m-%d")),
        (10, "47110005", "Ball bearing 6204-2RS (ab 500 Stk)", 500, "PCS", "4.35", 1,
         GUELTIG_AB.strftime("%Y-%m-%d")),
        (10, "47110005", "Ball bearing 6204-2RS (ab 1000 Stk)", 1000, "PCS", "4.10", 1,
         GUELTIG_AB.strftime("%Y-%m-%d")),
        (None, None, None, None, None, None, None, None),
        (20, "49900010", "Hydraulic hose DN12 2SN", 500, "M", "6.95", 1,
         GUELTIG_AB.strftime("%Y-%m-%d")),
        (30, "47119999", "Unknown part (nicht im Materialstamm)", 10, "PCS", "99.00", 1,
         GUELTIG_AB.strftime("%Y-%m-%d")),
        (None, None, "Subtotal", None, None, "3,952.50", None, None),
        (None, None, "VAT 19 %", None, None, "750.98", None, None),
        (None, None, "Total", None, None, "4,703.48", None, None),
    ]
    for index, zeile in enumerate(zeilen, start=7):
        for spalte, wert in enumerate(zeile, start=1):
            if wert is not None:
                blatt.cell(row=index, column=spalte, value=wert)
    mappe.save(pfad)


# ---------------------------------------------------------------------------
# 3: PDF
# ---------------------------------------------------------------------------

def pdf_angebot(pfad: Path) -> None:
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("  PyMuPDF fehlt – PDF-Beispiel wird uebersprungen.")
        return

    dokument = fitz.open()
    seite = dokument.new_page()

    def schreibe(x: float, y: float, text: str, groesse: float = 10,
                 fett: bool = False) -> None:
        seite.insert_text((x, y), text, fontsize=groesse,
                          fontname="helv" if not fett else "hebo")

    schreibe(50, 60, "Pumpen Weber GmbH & Co. KG", 16, True)
    schreibe(50, 78, "Werkstrasse 3  •  68199 Mannheim")
    schreibe(50, 92, "Telefon 0621 555-0  •  angebote@pumpen-weber.de")

    schreibe(50, 130, "ANGEBOT", 14, True)
    schreibe(50, 152, "Angebots-Nr.:  AG-2026-1188")
    schreibe(50, 168, f"Datum:  {_datum(HEUTE)}")
    schreibe(50, 184, f"Freibleibend gueltig bis:  {_datum(HEUTE + timedelta(days=60))}")
    schreibe(300, 152, "Kundennummer:  47110")
    schreibe(300, 168, "Zahlungsziel:  60 Tage netto")
    schreibe(300, 184, "Lieferbedingung:  CPT Werk")
    schreibe(300, 200, "Waehrung:  EUR")

    spalten = (50, 105, 175, 330, 385, 430, 490)
    ueberschriften = ("Pos", "Material", "Bezeichnung", "Menge", "ME", "Preis", "gueltig ab")
    for x, text in zip(spalten, ueberschriften):
        schreibe(x, 240, text, 10, True)
    seite.draw_line(fitz.Point(50, 245), fitz.Point(560, 245))

    zeilen = (
        ("10", "48200110", "Kreiselpumpe Typ KP-40", "12", "ST", "1.298,00",
         _datum(GUELTIG_AB)),
        ("20", "48200111", "Gleitringdichtung KP-40", "40", "ST", "289,00",
         _datum(GUELTIG_AB)),
        ("30", "47110003", "Wellendichtring FPM 30x47x7", "100", "ST", "19,80",
         _datum(GUELTIG_AB)),
    )
    y = 265
    for zeile in zeilen:
        for x, text in zip(spalten, zeile):
            schreibe(x, y, text)
        y += 20

    schreibe(50, y + 30, "Preise verstehen sich netto ab Werk zuzueglich gesetzlicher "
                         "Mehrwertsteuer.")
    schreibe(50, y + 46, "Mindestbestellmenge Pumpen: 5 Stueck.  Lieferzeit 6 Wochen.")
    schreibe(50, y + 78, "Mit freundlichen Gruessen")
    schreibe(50, y + 94, "M. Weber, Vertriebsleitung")

    dokument.save(pfad)
    dokument.close()


# ---------------------------------------------------------------------------
# 4 + 5: E-Mails
# ---------------------------------------------------------------------------

_MAIL_TEXT = f"""Sehr geehrte Damen und Herren,

vielen Dank fuer Ihre Anfrage. Aufgrund gestiegener Rohstoffpreise muessen wir
unsere Preise zum {_datum(GUELTIG_AB)} wie folgt anpassen:

  - Artikel 47110001, Dichtring NBR 40x52x7: 12,85 EUR je Stueck (bisher 12,40 EUR)
  - Artikel 47110002, O-Ring Viton 25x3: 0,89 EUR/St ab 500 Stueck
  - Artikel 47110003, Wellendichtring FPM 30x47x7: 18,95 EUR/St (unveraendert)

Die Mindestbestellmenge betraegt weiterhin 50 Stueck, die Lieferzeit 14 Tage.
Unsere Angebotsnummer lautet ANG-2026-04712, das Angebot ist 60 Tage gueltig.
Zahlungsbedingungen: 30 Tage netto.

Mit freundlichen Gruessen

Thomas Wagner
Vertrieb
Muster Dichtungstechnik GmbH
Industriestrasse 14, 33602 Bielefeld
Telefon 0521 555-120
vertrieb@muster-dichtungstechnik.de

Diese E-Mail enthaelt vertrauliche Informationen. Sollten Sie nicht der richtige
Adressat sein, informieren Sie bitte den Absender und loeschen Sie diese Mail.
"""

_MAIL_HTML = f"""<html><body>
<p>Guten Tag,</p>
<p>anbei unsere aktualisierten Preise, gueltig ab {_datum(GUELTIG_AB)}.
Unsere Angebotsnummer: <b>Q-2026-8842</b>.</p>
<table border="1" cellpadding="4">
  <tr><th>Material</th><th>Bezeichnung</th><th>Menge</th><th>ME</th>
      <th>Preis</th><th>PE</th></tr>
  <tr><td>47110005</td><td>Kugellager 6204-2RS</td><td>1.000</td><td>ST</td>
      <td>4,35 EUR</td><td>1</td></tr>
  <tr><td>49900010</td><td>Hydraulikschlauch DN12 2SN</td><td>500</td><td>M</td>
      <td>7,10 EUR</td><td>1</td></tr>
</table>
<p>Zahlungsbedingungen: 30 Tage netto. Lieferzeit 21 Tage.</p>
<p>Freundliche Gruesse<br>Nordtec Industriebedarf AG</p>
</body></html>"""


def mail_freitext(pfad: Path) -> None:
    nachricht = EmailMessage()
    nachricht["From"] = "Thomas Wagner <vertrieb@muster-dichtungstechnik.de>"
    nachricht["To"] = "einkauf@unsere-firma.de"
    nachricht["Subject"] = f"Preisanpassung zum {_datum(GUELTIG_AB)} – Angebot ANG-2026-04712"
    nachricht["Date"] = HEUTE.strftime("%a, %d %b %Y 09:14:00 +0200")
    nachricht.set_content(_MAIL_TEXT)
    pfad.write_bytes(nachricht.as_bytes())


def mail_mit_anhang(pfad: Path, anhang: Path) -> None:
    nachricht = EmailMessage()
    nachricht["From"] = "Vertrieb Nordtec <vertrieb@nordtec-industriebedarf.de>"
    nachricht["To"] = "einkauf@unsere-firma.de"
    nachricht["Subject"] = "Angebot Q-2026-8842 – Preisliste im Anhang"
    nachricht["Date"] = HEUTE.strftime("%a, %d %b %Y 11:02:00 +0200")
    nachricht.set_content("Guten Tag,\n\nanbei unsere aktuelle Preisliste.\n\n"
                          "Freundliche Gruesse\nNordtec Industriebedarf AG\n")
    nachricht.add_alternative(_MAIL_HTML, subtype="html")
    if anhang.exists():
        nachricht.add_attachment(
            anhang.read_bytes(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=anhang.name)
    pfad.write_bytes(nachricht.as_bytes())


# ---------------------------------------------------------------------------
# 6 + 7: Text und CSV
# ---------------------------------------------------------------------------

def textdatei(pfad: Path) -> None:
    pfad.write_text(_MAIL_TEXT, encoding="utf-8")


def csv_datei(pfad: Path) -> None:
    with pfad.open("w", encoding="utf-8-sig", newline="") as datei:
        schreiber = csv.writer(datei, delimiter=";")
        schreiber.writerow(["Pos", "Material", "Bezeichnung", "Menge", "ME", "Preis",
                            "PE", "Waehrung", "Gueltig ab"])
        schreiber.writerow([10, "47110001", "Dichtring NBR 40x52x7", "500", "ST",
                            "12,85", "1", "EUR", _datum(GUELTIG_AB)])
        schreiber.writerow([20, "47110002", "O-Ring Viton 25x3", "2.000", "ST",
                            "8,90", "10", "EUR", _datum(GUELTIG_AB)])
        schreiber.writerow([30, "49900011", "Schlauchschelle W2 20-32", "1.500", "ST",
                            "0,42", "1", "EUR", _datum(GUELTIG_AB)])


# ---------------------------------------------------------------------------
# 9: PDF im Fliesstextstil -- keine Tabelle, Preise stehen in Saetzen
# ---------------------------------------------------------------------------

_FLIESSTEXT_ABSAETZE = (
    "Dichtungswerk Sued GmbH",
    "Am Hafen 22  -  68159 Mannheim",
    "",
    "Sehr geehrte Damen und Herren,",
    "",
    "vielen Dank fuer Ihre Anfrage. Unser Angebot AG-2026-3355 vom "
    f"{_datum(HEUTE)} ist",
    "60 Tage bindend; das Zahlungsziel betraegt 30 Tage netto, alle Preise "
    "verstehen sich in EUR.",
    "",
    "Fuer den Dichtring NBR 40x52x7, Ihre Materialnummer 47110001, bieten wir "
    "Ihnen bei",
    "Abnahme von 500 Stueck einen Preis von 12,85 EUR je Stueck an. Die "
    "Mindestbestellmenge",
    "betraegt 50 Stueck, die Lieferzeit 14 Tage.",
    "",
    "Den O-Ring Viton 25x3, Ihre Materialnummer 47110002, liefern wir zu "
    "8,90 EUR je Stueck",
    "bei einer Mindestbestellmenge von 200 Stueck; die Lieferzeit betraegt "
    "ebenfalls 14 Tage.",
    "",
    "Fuer die Gleitringdichtung KP-40, Ihre Materialnummer 48200111, koennen "
    "wir derzeit",
    "keinen festen Preis nennen - dieser erfolgt auf Anfrage.",
    "",
    f"Die genannten Preise gelten ab {_datum(GUELTIG_AB)}.",
    "",
    "Mit freundlichen Gruessen",
    "K. Brandt, Vertrieb",
)


def pdf_fliesstext(pfad: Path) -> None:
    """Angebot ohne jede Tabelle -- alles steht im Fliesstext."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("  PyMuPDF fehlt - PDF-Beispiel wird uebersprungen.")
        return

    dokument = fitz.open()
    seite = dokument.new_page()
    y = 70
    for zeile in _FLIESSTEXT_ABSAETZE:
        if zeile:
            seite.insert_text((55, y), zeile, fontsize=10.5, fontname="helv")
        y += 16
    dokument.save(pfad)
    dokument.close()


# ---------------------------------------------------------------------------
# 10: Excel mit quer verteilten Kopfdaten
# ---------------------------------------------------------------------------

def excel_kopfdaten_quer(pfad: Path) -> None:
    """Kopfzeile erst in Zeile 12, Datum unter der Tabelle, Waehrung im Fuss."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = "Angebot"

    blatt["A1"] = "[ Firmenlogo ]"
    blatt["G2"] = "Angebot Nr.:"
    blatt["H2"] = "ANG-2026-7788"          # ganz rechts, weit weg vom Rest
    blatt["A4"] = "Sehr geehrte Damen und Herren,"
    blatt["A6"] = ("vielen Dank fuer Ihre Anfrage. Gerne unterbreiten wir Ihnen "
                   "folgendes Angebot.")
    blatt["A8"] = "Ansprechpartner:"
    blatt["B8"] = "Frau Sabine Groeger"
    blatt["A9"] = "Lieferbedingung:"
    blatt["B9"] = "FCA Bielefeld"

    kopf = ["Pos.", "Materialnummer", "Bezeichnung", "Menge", "ME", "Preis"]
    for spalte, text in enumerate(kopf, start=1):
        blatt.cell(row=12, column=spalte, value=text).font = Font(bold=True)

    zeilen = [
        (10, "47110001", "Öldichtring NBR 40x52x7, ölbeständig", 500, "Stk.", "12,85"),
        (20, "49900012", "Meßstab Edelstahl 1000 mm (±0,5 mm)", 120, "Meter", "6,40"),
    ]
    for index, zeile in enumerate(zeilen, start=13):
        for spalte, wert in enumerate(zeile, start=1):
            blatt.cell(row=index, column=spalte, value=wert)

    blatt["A15"] = "Angebotsdatum:"
    blatt["B15"] = _datum(HEUTE)           # unter der Tabelle, nicht darueber
    blatt["A17"] = "Zahlungsbedingungen:"
    blatt["B17"] = "30 Tage netto"
    blatt["A20"] = "Alle Preise in EUR zuzueglich gesetzlicher Mehrwertsteuer."

    mappe.save(pfad)


# ---------------------------------------------------------------------------
# 11: E-Mail, deren Text die entscheidenden Ergaenzungen traegt
# ---------------------------------------------------------------------------

_ERGAENZUNGS_CSV = (
    "Pos;Material;Bezeichnung;Menge;ME;Preis\n"
    "10;47110001;Dichtring NBR 40x52x7;500;St;12,85\n"
    "20;47110002;O-Ring Viton 25x3;200;St;8,90\n"
    "30;47110003;Wellendichtring FPM 30x47x7;100;St;18,95\n"
)

_ERGAENZUNGS_MAIL = f"""Sehr geehrte Damen und Herren,

anbei unsere Preisliste. Die Preise gelten ab {_datum(GUELTIG_AB)},
Zahlungsziel 30 Tage netto.

Position 30 entfaellt, dafuer neu: 47110009 zu 4,20 EUR.
Fuer Artikel 47110001 gilt abweichend eine Mindestmenge von 500 Stueck.
Der Preis fuer 47110002 im Anhang ist ueberholt, es gilt 9,10 EUR.

Mit freundlichen Gruessen
Thomas Wagner
Muster Dichtungstechnik GmbH
"""


def mail_mit_ergaenzungen(pfad: Path) -> None:
    """Der Alltagsfall: Tabelle im Anhang, das Wichtige im Mailtext."""
    nachricht = EmailMessage()
    nachricht["From"] = "Thomas Wagner <vertrieb@muster-dichtungstechnik.de>"
    nachricht["To"] = "einkauf@unsere-firma.de"
    nachricht["Subject"] = "Preisliste 2026 - bitte Ergaenzungen im Text beachten"
    nachricht["Date"] = HEUTE.strftime("%a, %d %b %Y 08:35:00 +0200")
    nachricht.set_content(_ERGAENZUNGS_MAIL)
    nachricht.add_attachment(_ERGAENZUNGS_CSV.encode("utf-8"), maintype="text",
                             subtype="csv", filename="Preisliste_2026.csv")
    pfad.write_bytes(nachricht.as_bytes())


# ---------------------------------------------------------------------------
# 12: PDF mit Staffelpreisen ueber zwei Seiten
# ---------------------------------------------------------------------------

def pdf_staffelpreise_zwei_seiten(pfad: Path) -> None:
    """Positionsliste laeuft ueber den Seitenumbruch -- inkl. Uebertragszeile."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("  PyMuPDF fehlt - PDF-Beispiel wird uebersprungen.")
        return

    spalten = (50, 105, 185, 350, 400, 460)
    ueberschriften = ("Pos", "Material", "Bezeichnung", "Menge", "ME", "Preis")

    dokument = fitz.open()

    def kopfzeile(seite, y: float) -> float:
        for x, text in zip(spalten, ueberschriften):
            seite.insert_text((x, y), text, fontsize=10, fontname="hebo")
        seite.draw_line(fitz.Point(50, y + 5), fitz.Point(540, y + 5))
        return y + 22

    def zeilen_schreiben(seite, y: float, zeilen) -> float:
        for zeile in zeilen:
            for x, text in zip(spalten, zeile):
                seite.insert_text((x, y), text, fontsize=10, fontname="helv")
            y += 18
        return y

    seite = dokument.new_page()
    seite.insert_text((50, 60), "Nordtec Industriebedarf AG", fontsize=15,
                      fontname="hebo")
    seite.insert_text((50, 82), "Angebots-Nr.:  AG-2026-9001", fontsize=10,
                      fontname="helv")
    seite.insert_text((50, 98), f"Datum:  {_datum(HEUTE)}", fontsize=10,
                      fontname="helv")
    seite.insert_text((50, 114), "Waehrung:  EUR   -   Zahlungsziel: 30 Tage netto",
                      fontsize=10, fontname="helv")
    seite.insert_text((50, 130), "Seite 1 von 2", fontsize=9, fontname="helv")

    y = kopfzeile(seite, 170)
    y = zeilen_schreiben(seite, y, (
        ("10", "47110005", "Kugellager 6204-2RS", "100", "Stk.", "4,75"),
        ("", "", "ab 500 Stk 4,35 EUR", "", "", ""),
        ("", "", "ab 1000 Stk 4,10 EUR", "", "", ""),
        ("20", "49900010", "Hydraulikschlauch DN12 2SN", "500", "Meter", "6,95"),
        ("", "", "ab 2000 m 6,40 EUR", "", "", ""),
    ))
    seite.insert_text((350, y + 14), "Uebertrag:  1.622,50 EUR", fontsize=10,
                      fontname="helv")

    seite = dokument.new_page()
    seite.insert_text((50, 60), "Seite 2 von 2", fontsize=9, fontname="helv")
    seite.insert_text((350, 60), "Uebertrag:  1.622,50 EUR", fontsize=10,
                      fontname="helv")
    y = kopfzeile(seite, 100)
    y = zeilen_schreiben(seite, y, (
        ("30", "48200111", "Gleitringdichtung KP-40", "40", "St", "289,00"),
        ("", "", "ab 100 St 265,00 EUR", "", "", ""),
        ("40", "47110004", "Flachdichtung Klingersil 100x60x2", "800", "PCS", "3,40"),
    ))
    seite.insert_text((350, y + 20), "Gesamtsumme:  15.402,50 EUR", fontsize=10,
                      fontname="helv")

    dokument.save(pfad)
    dokument.close()


# ---------------------------------------------------------------------------
# 13: Vertauschte Artikelnummernspalten, englische Ueberschriften
# ---------------------------------------------------------------------------

def excel_vertauschte_spalten(pfad: Path) -> None:
    """"Part No" ist die Nummer des Lieferanten, "Customer Material" unsere.

    Zusaetzlich stehen in derselben Datei beide Zahlenformate nebeneinander
    (1,234.56 englisch und 1.234,56 deutsch), es gibt Umlaute in den
    Bezeichnungen, drei Schreibweisen derselben Mengeneinheit und eine
    Position ganz ohne Preis ("auf Anfrage").
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = "Quotation"

    blatt["A1"] = "Global Sealing Solutions Ltd."
    blatt["A1"].font = Font(bold=True)
    blatt["A2"] = "Quotation No. GS-2026-0042"
    blatt["A3"] = f"Date: {HEUTE.strftime('%Y-%m-%d')}"
    blatt["A4"] = "Currency: EUR"

    kopf = ["Item", "Part No", "Customer Material", "Description", "Qty",
            "Unit", "Unit price"]
    for spalte, text in enumerate(kopf, start=1):
        blatt.cell(row=6, column=spalte, value=text).font = Font(bold=True)

    zeilen = [
        (10, "DR-40527-NBR", "47110001", "Öldichtring NBR 40x52x7", "1,234.56",
         "PCS", "12.85"),
        (20, "OR-2503-FPM", "47110002", "O-Ring Viton 25x3 (Übermaß)", "1.234,56",
         "Stk.", "8,90"),
        (30, "GLR-KP40", "48200111", "Gleitringdichtung KP-40, größere Bauform",
         "40", "St", "auf Anfrage"),
        (40, "HS-DN12", "49900010", "Hydraulikschlauch DN12 2SN", "500", "Meter",
         "6.95"),
    ]
    for index, zeile in enumerate(zeilen, start=7):
        for spalte, wert in enumerate(zeile, start=1):
            blatt.cell(row=index, column=spalte, value=wert)

    blatt["A12"] = "Payment terms: 30 days net.  Prices ex works."
    mappe.save(pfad)



# ---------------------------------------------------------------------------
# Textverarbeitungsformate: Word, OpenDocument, RTF
# ---------------------------------------------------------------------------

_W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _docx_absatz(text: str) -> str:
    return f'<w:p><w:r><w:t xml:space="preserve">{text}</w:t></w:r></w:p>'


def _docx_zeile(*werte: str) -> str:
    zellen = "".join(
        f'<w:tc><w:p><w:r><w:t xml:space="preserve">{w}</w:t></w:r></w:p></w:tc>'
        for w in werte)
    return f"<w:tr>{zellen}</w:tr>"


def word_angebot(pfad: Path) -> None:
    """Angebot als Word-Datei -- viele Lieferanten schicken genau das.

    Wird ohne Zusatzbibliothek erzeugt: .docx ist ein ZIP mit XML darin.
    """
    import zipfile

    koerper = (
        _docx_absatz("Schmidt &amp; Partner Werkstoffe KG")
        + _docx_absatz("Kronenstrasse 8, 42651 Solingen")
        + _docx_absatz("Angebot Nr. ANG-2026-7788 vom " + _datum(HEUTE))
        + _docx_absatz("Waehrung: EUR&#9;Zahlungsziel: 30 Tage netto")
        + _docx_absatz("Preise gueltig ab " + _datum(GUELTIG_AB))
        + "<w:tbl>"
        + _docx_zeile("Pos.", "Ihre Artikelnummer", "Unsere Art.-Nr.",
                      "Bezeichnung", "Menge", "ME", "Preis")
        + _docx_zeile("10", "47110001", "SP-DR-4052", "Dichtring NBR 40x52x7",
                      "500", "Stk.", "12,95")
        + _docx_zeile("20", "47110005", "SP-KL-6204", "Kugellager 6204-2RS",
                      "1.000", "St", "4,45")
        + _docx_zeile("30", "49900010", "SP-HS-DN12", "Hydraulikschlauch DN12 2SN",
                      "300", "Meter", "7,30")
        + _docx_zeile("40", "47110004", "SP-FD-1006", "Flachdichtung 100x60x2",
                      "800", "Stk.", "auf Anfrage")
        + "</w:tbl>"
        + _docx_absatz("Mindestbestellmenge 50 Stueck, Lieferzeit 14 Tage.")
        + _docx_absatz("Das Angebot ist 60 Tage gueltig.")
    )
    dokument = (f'<?xml version="1.0" encoding="UTF-8"?>'
                f'<w:document xmlns:w="{_W_NS}"><w:body>{koerper}</w:body></w:document>')
    beziehungen = (
        '<?xml version="1.0" encoding="UTF-8"?><Relationships '
        'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/></Relationships>')
    typen = (
        '<?xml version="1.0" encoding="UTF-8"?><Types '
        'xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Default Extension="rels" ContentType="application/vnd.'
        'openxmlformats-package.relationships+xml"/></Types>')

    with zipfile.ZipFile(pfad, "w", zipfile.ZIP_DEFLATED) as archiv:
        archiv.writestr("[Content_Types].xml", typen)
        archiv.writestr("_rels/.rels", beziehungen)
        archiv.writestr("word/document.xml", dokument)


def opendocument_angebot(pfad: Path) -> None:
    """Dasselbe als OpenDocument (LibreOffice)."""
    import zipfile

    T = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    TB = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    O = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"

    def zelle(text: str) -> str:
        return f'<table:table-cell><text:p>{text}</text:p></table:table-cell>'

    def zeile(*werte: str) -> str:
        return "<table:table-row>" + "".join(zelle(w) for w in werte) + "</table:table-row>"

    inhalt = (
        f'<?xml version="1.0" encoding="UTF-8"?>'
        f'<office:document-content xmlns:office="{O}" xmlns:text="{T}" '
        f'xmlns:table="{TB}"><office:body><office:text>'
        f'<text:p>Nordtec Industriebedarf AG</text:p>'
        f'<text:p>Quotation No. Q-2026-9911, date {_datum(HEUTE)}</text:p>'
        f'<text:p>Currency: EUR, payment terms 30 days net</text:p>'
        f'<table:table>'
        + zeile("Item", "Customer Material", "Description", "Qty", "Unit", "Unit price")
        + zeile("10", "47110005", "Ball bearing 6204-2RS", "1,000", "PCS", "4.55")
        + zeile("20", "49900010", "Hydraulic hose DN12", "500", "M", "7.15")
        + f'</table:table>'
        f'<text:p>Prices valid from {_datum(GUELTIG_AB)}.</text:p>'
        f'</office:text></office:body></office:document-content>')

    with zipfile.ZipFile(pfad, "w", zipfile.ZIP_DEFLATED) as archiv:
        archiv.writestr("mimetype", "application/vnd.oasis.opendocument.text")
        archiv.writestr("content.xml", inhalt)


def rtf_angebot(pfad: Path) -> None:
    """Kurzes RTF-Angebot (kommt selten, aber es kommt vor).

    RTF kennt keine Tabellenstruktur -- Zellen und Zeilen sind Steuerworte
    (cell, row).  Genau daraus baut der Leser die Tabelle wieder zusammen.
    """
    s = chr(92)      # Rueckstrich, damit die Quelltextzeilen lesbar bleiben

    def tabellenzeile(*werte: str) -> str:
        return (s + "cell ").join(werte) + s + "cell" + s + "row"

    zeilen = [
        "{" + s + "rtf1" + s + "ansi" + s + "deff0",
        "{" + s + "fonttbl{" + s + "f0 Arial;}}",
        "Pumpen Weber GmbH " + s + "& Co. KG" + s + "par",
        f"Angebot Nr. AG-2026-4455 vom {_datum(HEUTE)}" + s + "par",
        "Waehrung EUR, Zahlungsziel 60 Tage netto" + s + "par",
        tabellenzeile("Pos.", "Material", "Bezeichnung", "Menge", "ME", "Preis"),
        tabellenzeile("10", "48200110", "Kreiselpumpe KP-40", "12", "ST", "1.310,00"),
        tabellenzeile("20", "48200111", "Gleitringdichtung KP-40", "40", "ST", "292,00"),
        s + "par Preise gueltig ab " + _datum(GUELTIG_AB) + s + "par",
        "}",
    ]
    pfad.write_text(chr(10).join(zeilen), encoding="cp1252")


# ---------------------------------------------------------------------------
# OpenDocument-Tabelle (.ods) und ZIP-Archiv
# ---------------------------------------------------------------------------

def ods_preisliste(pfad: Path) -> None:
    """LibreOffice-Calc-Angebot mit zwei Blaettern.

    Blatt 1 traegt Anschreiben und Kopfdaten, Blatt 2 die Positionen -- genau
    so, wie Lieferanten ihre Preislisten tatsaechlich aufbauen.

    Bewusst eingebaut: zusammengefasste Leerspalten
    (``table:number-columns-repeated``) und zusammengefasste Leerzeilen
    (``table:number-rows-repeated``).  So speichert LibreOffice wirklich --
    und genau daran scheitert ein naiver Leser.  Ebenfalls bewusst: die
    Preise stehen als roher ``office:value`` in der Datei, waehrend der
    Anzeigetext deutsch formatiert ist ("4,55 EUR").
    """
    import zipfile

    T = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    TB = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    O = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"

    def text_zelle(inhalt: str, wiederholt: int = 1) -> str:
        wdh = f' table:number-columns-repeated="{wiederholt}"' if wiederholt > 1 else ""
        if not inhalt:
            return f"<table:table-cell{wdh}/>"
        return (f'<table:table-cell{wdh} office:value-type="string">'
                f"<text:p>{inhalt}</text:p></table:table-cell>")

    def zahl_zelle(wert: str, anzeige: str) -> str:
        """Rohwert und (abweichender) formatierter Anzeigetext."""
        return (f'<table:table-cell office:value-type="float" office:value="{wert}">'
                f"<text:p>{anzeige}</text:p></table:table-cell>")

    def zeile(*zellen: str, wiederholt: int = 1) -> str:
        wdh = f' table:number-rows-repeated="{wiederholt}"' if wiederholt > 1 else ""
        # Wie LibreOffice: die Zeile wird mit zusammengefassten Leerspalten
        # bis zum Blattrand aufgefuellt.
        fuellung = text_zelle("", 1014)
        return (f"<table:table-row{wdh}>" + "".join(zellen) + fuellung
                + "</table:table-row>")

    kopf = (
        '<table:table table:name="Anschreiben">'
        + zeile(text_zelle("Nordtec Industriebedarf AG"))
        + zeile(text_zelle("Angebot Nr."), text_zelle("ANG-2026-5501"))
        + zeile(text_zelle("Datum"), text_zelle(_datum(HEUTE)))
        + zeile(text_zelle("Waehrung"), text_zelle("EUR"))
        + zeile(text_zelle("Zahlungsziel"), text_zelle("30 Tage netto"))
        + zeile(text_zelle(""), wiederholt=1048570)
        + "</table:table>"
    )

    positionen = (
        ("10", "47110001", "Dichtring 52x72x10 NBR", "25", "Stk.", "12.35", "12,35 EUR"),
        ("20", "47110005", "Rillenkugellager 6204-2RS", "1000", "ST", "4.55", "4,55 EUR"),
        ("30", "49900010", "Hydraulikschlauch DN12", "500", "Meter", "7.15", "7,15 EUR"),
    )
    zeilen = [zeile(
        text_zelle("Pos"), text_zelle("Material"), text_zelle("Bezeichnung"),
        text_zelle("Menge"), text_zelle("ME"), text_zelle("Preis"))]
    for pos, material, bezeichnung, menge, me, wert, anzeige in positionen:
        zeilen.append(zeile(
            text_zelle(pos), text_zelle(material), text_zelle(bezeichnung),
            text_zelle(menge), text_zelle(me), zahl_zelle(wert, anzeige)))
    zeilen.append(zeile(text_zelle(""), wiederholt=1048000))

    tabelle = ('<table:table table:name="Positionen">' + "".join(zeilen)
               + "</table:table>")

    inhalt = (
        f'<?xml version="1.0" encoding="UTF-8"?>'
        f'<office:document-content xmlns:office="{O}" xmlns:text="{T}" '
        f'xmlns:table="{TB}"><office:body><office:spreadsheet>'
        f"{kopf}{tabelle}"
        f"</office:spreadsheet></office:body></office:document-content>")

    with zipfile.ZipFile(pfad, "w", zipfile.ZIP_DEFLATED) as archiv:
        archiv.writestr("mimetype",
                        "application/vnd.oasis.opendocument.spreadsheet")
        archiv.writestr("content.xml", inhalt)


def zip_sammlung(pfad: Path, *quellen: Path) -> None:
    """ZIP mit PDF + Excel -- und einer Datei, die der Import nicht kennt.

    Die .png-Attrappe ist Absicht: Lieferanten packen regelmaessig ihr Logo
    oder einen Scan mit ins Archiv.  Der Import muss so etwas sichtbar
    ueberspringen, statt es zu deuten.
    """
    import zipfile

    with zipfile.ZipFile(pfad, "w", zipfile.ZIP_DEFLATED) as archiv:
        for quelle in quellen:
            if quelle.exists():
                archiv.write(quelle, arcname=quelle.name)
        archiv.writestr("Firmenlogo.png", b"\x89PNG\r\n\x1a\n" + b"\x00" * 128)


# ---------------------------------------------------------------------------

def main() -> int:
    ZIEL.mkdir(parents=True, exist_ok=True)
    print(f"Beispieldateien werden erzeugt in: {ZIEL}")

    aufgaben = [
        ("Angebot_Muster_Dichtungstechnik.xlsx", excel_mit_kopfzeile),
        ("Preisliste_ohne_Kopfzeile.xlsx", excel_ohne_kopfzeile),
        ("Quotation_Nordtec_mit_Stoerfaellen.xlsx", excel_mit_stoerfaellen),
        ("Angebot_Pumpen_Weber.pdf", pdf_angebot),
        ("Preisanpassung_Muster.eml", mail_freitext),
        ("Preismitteilung.txt", textdatei),
        ("Preisliste_Muster.csv", csv_datei),
        ("Angebot_Schmidt_Partner.docx", word_angebot),
        ("Quotation_Nordtec.odt", opendocument_angebot),
        ("Angebot_Pumpen_Weber.rtf", rtf_angebot),
        # -- die schwierigen Faelle --------------------------------------
        ("Angebot_Fliesstext_Dichtungswerk.pdf", pdf_fliesstext),
        ("Angebot_Kopfdaten_quer_verteilt.xlsx", excel_kopfdaten_quer),
        ("Mail_mit_Ergaenzungen_im_Text.eml", mail_mit_ergaenzungen),
        ("Angebot_Staffelpreise_zwei_Seiten.pdf", pdf_staffelpreise_zwei_seiten),
        ("Quotation_vertauschte_Spalten.xlsx", excel_vertauschte_spalten),
        ("Preisliste_Nordtec.ods", ods_preisliste),
    ]
    # Hinweis: Die Windows-Konsole nutzt cp1252 -- deshalb bewusst nur ASCII
    # ausgeben, sonst bricht die Ausgabe mit einem UnicodeEncodeError ab.
    for name, funktion in aufgaben:
        ziel = ZIEL / name
        try:
            funktion(ziel)
            print(f"  [ok]     {name}")
        except Exception as fehler:  # noqa: BLE001 - Beispiele duerfen nie den Rest stoppen
            print(f"  [FEHLER] {name}: {fehler}")

    try:
        mail_mit_anhang(ZIEL / "Angebot_Nordtec_mit_Anhang.eml",
                        ZIEL / "Quotation_Nordtec_mit_Stoerfaellen.xlsx")
        print("  [ok]     Angebot_Nordtec_mit_Anhang.eml")
    except Exception as fehler:  # noqa: BLE001
        print(f"  [FEHLER] Angebot_Nordtec_mit_Anhang.eml: {fehler}")

    # Das ZIP verpackt bereits erzeugte Beispiele -- deshalb erst zum Schluss.
    try:
        zip_sammlung(ZIEL / "Angebot_Sammlung.zip",
                     ZIEL / "Angebot_Pumpen_Weber.pdf",
                     ZIEL / "Angebot_Muster_Dichtungstechnik.xlsx")
        print("  [ok]     Angebot_Sammlung.zip")
    except Exception as fehler:  # noqa: BLE001
        print(f"  [FEHLER] Angebot_Sammlung.zip: {fehler}")

    print("\nFertig. Ziehen Sie eine dieser Dateien einfach in das Anwendungsfenster.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
