"""Leser fuer Tabellendateien: xlsx/xlsm (openpyxl) und csv/txt-Tabellen.

Besonderheiten aus der Praxis:

* Lieferanten liefern gern *mehrere* Arbeitsblaetter (Deckblatt + Preisliste),
  deshalb wird jedes Blatt als eigener :class:`TableBlock` gelesen.
* Verbundene Zellen ("Preisliste 2026" ueber fuenf Spalten) werden aufgeloest,
  damit die Kopfzeilenerkennung nicht ins Leere laeuft.
* CSV kommt mit allen denkbaren Trennzeichen und Kodierungen -- beides wird
  erkannt, nicht angenommen.
* ``.xls`` (altes Binaerformat) braucht ``xlrd``.  Fehlt das Paket, gibt es
  eine klare, freundliche Meldung statt eines Absturzes.
"""

from __future__ import annotations

import csv
import datetime as _dt
import logging
import re
from collections import Counter
from decimal import Decimal
from pathlib import Path

from ...models.enums import SourceKind
from ...utils.parsing import format_date
from .base import DocumentReader, RawDocument, TableBlock

logger = logging.getLogger(__name__)

__all__ = ["ExcelReader", "CsvReader", "XLS_HINT", "cell_to_text",
           "detect_delimiter", "repair_decimal_split_rows",
           "AMBIGUOUS_COMMA_HINT"]

AMBIGUOUS_COMMA_HINT = (
    "Das Komma ist in dieser Datei zugleich Feldtrenner und Dezimalzeichen. "
    "Das Trennzeichen ist damit nicht sicher bestimmbar")

XLS_HINT = ("Das alte Excel-Format .xls kann nicht gelesen werden (Paket 'xlrd' "
            "fehlt).  Bitte die Datei in Excel als .xlsx speichern -- oder "
            "'pip install xlrd' ausfuehren.")

#: Kodierungen in der Reihenfolge, in der sie probiert werden
_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")

#: Trennzeichen-Kandidaten, wenn der Sniffer versagt
_DELIMITERS = (";", ",", "\t", "|")

#: Maximale Zeilenzahl je Blatt (Schutz vor Riesendateien)
_MAX_ROWS = 20_000


def cell_to_text(value: object) -> str:
    """Excel-Zellwert in Text wandeln, ohne Information zu verlieren.

    Zahlen behalten ihre Nachkommastellen (kein ``1.2000000000000002``),
    Datumswerte werden deutsch formatiert, ``None`` wird zu ``""``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "ja" if value else "nein"
    if isinstance(value, _dt.datetime):
        if value.time() == _dt.time(0, 0):
            return format_date(value.date())
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, _dt.date):
        return format_date(value)
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        # Repraesentationsrauschen von Gleitkomma abschneiden
        text = f"{value:.10f}".rstrip("0").rstrip(".")
        return text or "0"
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


class ExcelReader(DocumentReader):
    """xlsx/xlsm ueber openpyxl, xls mit klarer Fehlermeldung."""

    extensions = (".xlsx", ".xlsm", ".xltx", ".xls")

    def read(self, path: str) -> RawDocument:
        document = RawDocument(source_path=str(path), source_kind=SourceKind.EXCEL)
        suffix = Path(path).suffix.lower()
        if suffix == ".xls":
            return self._read_legacy(path, document)

        try:
            from openpyxl import load_workbook
        except ImportError:
            document.add_warning("openpyxl ist nicht installiert -- Excel-Dateien "
                                 "koennen nicht gelesen werden.")
            return document

        try:
            workbook = load_workbook(filename=path, data_only=True, read_only=False)
        except Exception as exc:  # noqa: BLE001
            document.add_warning(f"Excel-Datei konnte nicht geoeffnet werden: {exc}")
            logger.warning("XLSX nicht lesbar (%s): %s", path, exc)
            return document

        try:
            self._read_workbook(workbook, document)
        except Exception as exc:  # noqa: BLE001
            document.add_warning(f"Excel-Datei nur teilweise lesbar: {exc}")
            logger.warning("XLSX-Auswertung unvollstaendig (%s): %s", path, exc, exc_info=True)
        finally:
            try:
                workbook.close()
            except Exception:  # noqa: BLE001
                pass

        document.text = "\n\n".join(table.as_text() for table in document.tables)
        logger.info("Excel gelesen: %s (%d Blaetter)", Path(path).name, len(document.tables))
        return document

    # ------------------------------------------------------------------
    def _read_workbook(self, workbook, document: RawDocument) -> None:
        document.meta["sheet_names"] = list(workbook.sheetnames)
        for sheet in workbook.worksheets:
            try:
                rows, merged = self._read_sheet(sheet)
            except Exception as exc:  # noqa: BLE001
                document.add_warning(f"Arbeitsblatt '{sheet.title}' nicht lesbar: {exc}")
                continue
            if not rows:
                continue
            block = TableBlock(rows=rows, page=0, origin="excel", title=str(sheet.title),
                               merged_cells=merged)
            document.tables.append(block)
            document.pages.append(block.as_text())

    def _read_sheet(self, sheet) -> tuple[list[list[str]], set[tuple[int, int]]]:
        """Ein Arbeitsblatt als Textmatrix; verbundene Zellen aufgeloest.

        Rueckgabe: (Zeilen, Koordinaten der aus Verbuenden gefuellten Zellen).
        """
        rows: list[list[str]] = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= _MAX_ROWS:
                logger.warning("Arbeitsblatt '%s' bei %d Zeilen abgeschnitten",
                               sheet.title, _MAX_ROWS)
                break
            rows.append([cell_to_text(value) for value in row])
        merged = self._fill_merged(sheet, rows)
        # Leere Zeilen am Ende entfernen
        while rows and not any(cell for cell in rows[-1]):
            rows.pop()
        return rows, merged

    @staticmethod
    def _fill_merged(sheet, rows: list[list[str]]) -> set[tuple[int, int]]:
        """Wert einer verbundenen Zelle in alle beteiligten Zellen kopieren.

        Rueckgabe: die (Zeile, Spalte)-Koordinaten aller Zellen, die dabei
        *gefuellt* wurden -- die Extraktion markiert deren Werte als unsicher.
        """
        filled: set[tuple[int, int]] = set()
        try:
            ranges = list(sheet.merged_cells.ranges)
        except Exception:  # noqa: BLE001 -- read_only-Modus kennt das nicht
            return filled
        for cell_range in ranges:
            try:
                min_row, min_col = cell_range.min_row, cell_range.min_col
                max_row, max_col = cell_range.max_row, cell_range.max_col
                if min_row - 1 >= len(rows):
                    continue
                source_row = rows[min_row - 1]
                if min_col - 1 >= len(source_row):
                    continue
                value = source_row[min_col - 1]
                if not value:
                    continue
                for r in range(min_row - 1, min(max_row, len(rows))):
                    row = rows[r]
                    for c in range(min_col - 1, min(max_col, len(row))):
                        if not row[c]:
                            row[c] = value
                            if (r, c) != (min_row - 1, min_col - 1):
                                filled.add((r, c))
            except Exception:  # noqa: BLE001
                continue
        return filled

    # ------------------------------------------------------------------
    def _read_legacy(self, path: str, document: RawDocument) -> RawDocument:
        try:
            import xlrd  # type: ignore
        except ImportError:
            document.add_warning(XLS_HINT)
            document.meta["needs_conversion"] = True
            logger.info("XLS ohne xlrd angefordert: %s", path)
            return document

        try:
            book = xlrd.open_workbook(path)
            for sheet in book.sheets():
                rows = [[cell_to_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                        for r in range(min(sheet.nrows, _MAX_ROWS))]
                if rows:
                    document.tables.append(
                        TableBlock(rows=rows, origin="excel", title=str(sheet.name))
                    )
        except Exception as exc:  # noqa: BLE001
            document.add_warning(f"XLS-Datei nicht lesbar: {exc}")
        document.text = "\n\n".join(table.as_text() for table in document.tables)
        return document


class CsvReader(DocumentReader):
    """CSV/TSV mit automatischer Trennzeichen- und Kodierungserkennung."""

    extensions = (".csv", ".tsv")

    def read(self, path: str) -> RawDocument:
        document = RawDocument(source_path=str(path), source_kind=SourceKind.EXCEL)
        raw, encoding, warning = _read_text_file(path)
        if warning:
            document.add_warning(warning)
        if raw is None:
            return document
        document.meta["encoding"] = encoding

        delimiter = detect_delimiter(raw)
        document.meta["delimiter"] = delimiter
        try:
            reader = csv.reader(raw.splitlines(), delimiter=delimiter)
            rows = [[cell.strip() for cell in row] for row in reader]
        except csv.Error as exc:
            document.add_warning(f"CSV-Datei konnte nicht zerlegt werden: {exc}")
            document.text = raw
            return document

        if delimiter == ",":
            rows, geheilt, offen = repair_decimal_split_rows(rows)
            if geheilt:
                document.add_warning(
                    f"{AMBIGUOUS_COMMA_HINT}. In {geheilt} Zeile(n) wurden "
                    "zerrissene Betraege wieder zusammengefuegt (z. B. '2' und "
                    "'95' zu '2,95') -- bitte die Preise stichprobenartig "
                    "pruefen.")
            if offen:
                document.add_warning(
                    f"{AMBIGUOUS_COMMA_HINT}. In {offen} Zeile(n) laesst sich "
                    "nicht eindeutig entscheiden, welche Kommas Feldtrenner "
                    "und welche Dezimalzeichen sind -- diese Zeilen wurden "
                    "unveraendert uebernommen und sind zu pruefen.")
            document.meta["comma_repaired_rows"] = geheilt
            document.meta["comma_ambiguous_rows"] = offen

        rows = [row for row in rows if any(cell for cell in row)]
        if rows:
            document.tables.append(
                TableBlock(rows=rows, origin="excel", title=Path(path).stem)
            )
        document.text = raw
        logger.info("CSV gelesen: %s (Trennzeichen %r, Kodierung %s, %d Zeilen)",
                    Path(path).name, delimiter, encoding, len(rows))
        return document


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def _read_text_file(path: str) -> tuple[str | None, str, str]:
    """Textdatei lesen und dabei die Kodierung bestimmen."""
    try:
        data = Path(path).read_bytes()
    except OSError as exc:
        return None, "", f"Datei konnte nicht gelesen werden: {exc}"
    for encoding in _ENCODINGS:
        try:
            return data.decode(encoding), encoding, ""
        except UnicodeDecodeError:
            continue
    return (data.decode("latin-1", "replace"), "latin-1",
            "Kodierung konnte nicht sicher bestimmt werden -- Umlaute pruefen.")


# ---------------------------------------------------------------------------
# Komma als Trenner UND als Dezimalzeichen
# ---------------------------------------------------------------------------
#
# Eine deutsche Preisliste mit Komma als Feldtrenner ist in sich
# widerspruechlich: "Dichtring,2,95" enthaelt drei Felder, gemeint sind zwei.
# Das Trennzeichen selbst ist dabei richtig erkannt -- zerrissen wird die
# *Zahl*.  Deshalb wird hier nicht am Trennzeichen gedreht, sondern die
# Zerlegung bewertet: hat eine Zeile mehr Felder als die Tabelle breit ist,
# und laesst sich der Ueberschuss durch genau *eine* moegliche Zusammenfassung
# von Zahlenbruchstuecken aufloesen, dann wird sie zusammengefuegt.  Gibt es
# mehrere Moeglichkeiten, wird NICHTS geraten -- es entsteht ein Befund.

#: Zweite Haelfte eines zerrissenen Betrags: die Cent-Stellen.
#:
#: Verlangt werden GENAU ZWEI Ziffern ("95", "00", "50") -- oder ein bis vier
#: Ziffern, wenn eine Waehrung oder Einheit dahintersteht ("5 EUR", "50 kg").
#: Diese Enge ist Absicht: eine einzelne Ziffer wuerde jede Rabattspalte
#: ("...,50,0,...") als zerrissenen Betrag deuten und die Nachbarspalte
#: verschlucken, vier Ziffern jede zweite Mengenspalte.  Preise mit drei oder
#: vier Nachkommastellen werden dadurch nicht mehr zusammengefuegt, sondern
#: als nicht aufloesbar gemeldet -- lieber ein Befund als ein falscher Wert.
#: Ein Punkt darf nicht vorkommen: "250.00" waere die englische Schreibweise
#: und damit ein ganz anderer Fall.
_DECIMAL_TAIL_RE = re.compile(
    r"^(?:\d{2}|\d{1,4}\s*[A-Za-z%€$£][A-Za-z%/€$£]{0,9}\.?)$")

#: Erste Haelfte: eine ganze Zahl, gern mit Punkt als Tausendertrennung
#: ("2", "127", "1.250"), optional mit vorangestelltem Waehrungszeichen.
_DECIMAL_HEAD_RE = re.compile(
    r"^(?:[€$£]|EUR|USD|CHF|GBP)?\s*[+-]?\d{1,3}(?:\.\d{3})*$", re.I)


def _decimal_junctions(fields: list[str]) -> list[int]:
    """Stellen, an denen zwei Nachbarfelder ein zerrissener Betrag sein koennen.

    Ein Index ``i`` bedeutet: ``fields[i]`` und ``fields[i + 1]`` ergaeben
    zusammen mit einem Komma wieder eine Zahl ("127" + "00 EUR").
    """
    stellen: list[int] = []
    for index in range(len(fields) - 1):
        kopf = fields[index].strip()
        rumpf = fields[index + 1].strip()
        if _DECIMAL_HEAD_RE.match(kopf) and _DECIMAL_TAIL_RE.match(rumpf):
            stellen.append(index)
    return stellen


def _merge_junctions(fields: list[str], stellen: list[int]) -> list[str]:
    """Die genannten Stellen wieder mit einem Komma zusammenfuegen."""
    zusammen = set(stellen)
    out: list[str] = []
    index = 0
    while index < len(fields):
        if index in zusammen and index + 1 < len(fields):
            out.append(f"{fields[index].strip()},{fields[index + 1].strip()}")
            index += 2
        else:
            out.append(fields[index])
            index += 1
    return out


def repair_decimal_split_rows(
        rows: list[list[str]]) -> tuple[list[list[str]], int, int]:
    """Am Dezimalkomma zerrissene Zeilen heilen.

    Rueckgabe: ``(Zeilen, geheilt, nicht_aufloesbar)``.

    Massgeblich ist die Spaltenzahl, die sich ergibt, wenn man in jeder Zeile
    alle Zahlenbruchstuecke zusammenfasst -- die haeufigste dieser Zahlen ist
    die Breite der Tabelle.  Eine Zeile wird nur dann angefasst, wenn sie zu
    *viele* Felder hat und der Ueberschuss genau den gefundenen Zusammenfass-
    Stellen entspricht.  Stehen zwei moegliche Stellen direkt nebeneinander
    ("1,1,1"), ist die Deutung nicht eindeutig -- dann bleibt die Zeile, wie
    sie ist, und der Aufrufer meldet das.
    """
    inhalt = [row for row in rows if any(cell.strip() for cell in row)]
    if len(inhalt) < 2:
        return rows, 0, 0

    breiten = [len(row) for row in inhalt]
    if len(set(breiten)) == 1:
        # Alle Zeilen gleich breit: es gibt keinen Ueberschuss, den man
        # zuordnen koennte.  Hier zu raten hiesse, eine heile Tabelle
        # kaputtzumachen.
        return rows, 0, 0

    stellen_je_zeile = {id(row): _decimal_junctions(row) for row in inhalt}
    kandidaten = Counter(len(row) - len(stellen_je_zeile[id(row)])
                         for row in inhalt)
    ziel = kandidaten.most_common(1)[0][0]
    if ziel < 2:
        return rows, 0, 0

    # 1. Durchgang: alle Zeilen heilen, bei denen die Deutung eindeutig ist --
    #    der Ueberschuss entspricht genau den gefundenen Stellen, und keine
    #    zwei davon stossen aneinander.
    entschieden: dict[int, list[int]] = {}
    gelernt: Counter[int] = Counter()
    for row in inhalt:
        stellen = stellen_je_zeile[id(row)]
        noetig = len(row) - ziel
        if noetig <= 0:
            continue
        if len(stellen) == noetig and not _stossen_aneinander(stellen):
            entschieden[id(row)] = stellen
            gelernt.update(stellen)

    # 2. Durchgang: die uebrigen Zeilen haben MEHR moegliche Stellen als
    #    noetig ("0,12,40,..." -- sowohl "0,12" als auch "12,40" saehen aus
    #    wie ein Betrag).  Statt zu raten wird verglichen: an welchen Stellen
    #    lag der Betrag in den Zeilen, die eindeutig waren?  Eine Tabelle hat
    #    ihre Preisspalte in jeder Zeile an derselben Stelle.  Nur wenn die
    #    gelernten Stellen den Ueberschuss genau erklaeren, wird zusammen-
    #    gefuegt -- sonst bleibt die Zeile unangetastet und wird gemeldet.
    haeufig = {stelle for stelle, anzahl in gelernt.items()
               if anzahl >= 2 or anzahl == len(entschieden)}
    for row in inhalt:
        if id(row) in entschieden:
            continue
        noetig = len(row) - ziel
        if noetig <= 0:
            continue
        passend = sorted(set(stellen_je_zeile[id(row)]) & haeufig)
        if len(passend) == noetig and not _stossen_aneinander(passend):
            entschieden[id(row)] = passend

    geheilt = 0
    offen = 0
    ergebnis: list[list[str]] = []
    for row in rows:
        stellen = entschieden.get(id(row))
        if stellen is not None:
            ergebnis.append(_merge_junctions(row, stellen))
            geheilt += 1
            continue
        ergebnis.append(row)
        if id(row) in stellen_je_zeile and len(row) > ziel:
            offen += 1
    return ergebnis, geheilt, offen


def _stossen_aneinander(stellen: list[int]) -> bool:
    """Zwei Zusammenfass-Stellen ueberlappen sich ("1,1,1") -- nicht eindeutig."""
    vorhanden = set(stellen)
    return any(stelle + 1 in vorhanden for stelle in stellen)


def _split_lines(lines: list[str], delimiter: str) -> list[list[str]]:
    """Zeilen mit einem Kandidaten zerlegen -- mit Ruecksicht auf Anfuehrungszeichen.

    Es wird bewusst der echte CSV-Leser benutzt und nicht ``str.count``:
    ein Trennzeichen INNERHALB von Anfuehrungszeichen ("Dichtring, gross")
    ist keins, und genau daran scheitert reines Zeichenzaehlen.
    """
    try:
        return [row for row in csv.reader(lines, delimiter=delimiter)]
    except csv.Error:
        return [line.split(delimiter) for line in lines]


def _table_score(lines: list[str], delimiter: str) -> tuple[float, int, int]:
    """Wie sehr sieht die Datei mit diesem Trennzeichen nach Tabelle aus?

    Rueckgabe: ``(Bewertung, Laenge des besten Blocks, Spaltenzahl)``.

    Massstab ist der laengste *zusammenhaengende* Block von Zeilen mit
    derselben Feldzahl.  Das ist der entscheidende Unterschied zum blossen
    Zaehlen: eine Tabelle besteht aus gleichfoermigen Zeilen, Fliesstext nicht.
    Die Spaltenzahl geht mit ein, damit ein Block aus sieben Spalten einen
    Block aus zwei Spalten schlaegt -- deutsche Prosa unter der Tabelle
    ("... innerhalb 14 Tagen, 30 Tage netto.") ergibt naemlich sehr wohl
    mehrere aufeinanderfolgende Zeilen mit je zwei Komma-"Feldern".
    """
    rows = _split_lines(lines, delimiter)
    bester_block, beste_breite = 0, 0
    lauf_breite, lauf_laenge = 0, 0
    for row in rows:
        breite = len(row)
        if breite < 2:
            lauf_breite, lauf_laenge = 0, 0
            continue
        if breite == lauf_breite:
            lauf_laenge += 1
        else:
            lauf_breite, lauf_laenge = breite, 1
        if (lauf_laenge * (lauf_breite - 1)) > (bester_block * max(beste_breite - 1, 1)):
            bester_block, beste_breite = lauf_laenge, lauf_breite
    if bester_block < 2:
        return 0.0, bester_block, beste_breite
    return float(bester_block * (beste_breite - 1)), bester_block, beste_breite


def delimiter_scores(sample: str) -> dict[str, float]:
    """Bewertung aller Trennzeichen-Kandidaten (hoeher ist besser)."""
    lines = [line for line in sample.splitlines()[:40] if line.strip()]
    return {candidate: _table_score(lines, candidate)[0]
            for candidate in _DELIMITERS}


def detect_delimiter(sample: str) -> str:
    """Trennzeichen einer CSV bestimmen.

    Frueher wurde gezaehlt, wie oft ein Zeichen je Zeile vorkommt.  Das
    verliert an zwei Stellen, die in echten Angeboten die Regel sind:

    * Unter der Tabelle steht Fliesstext ("Zahlungsbedingungen: 2 % Skonto
      bei Zahlung innerhalb 14 Tagen, 30 Tage netto.").  Solche Saetze
      enthalten Kommas, aber keine Semikolons -- beim reinen Zaehlen gewann
      dadurch das Komma, die Tabelle blieb ungetrennt und "2,95" zerfiel
      obendrein.  Der Beleg ergab null Positionen.
    * Ein Briefkopf ueber der Tabelle wirkte genauso.

    Deshalb wird jetzt *probeweise zerlegt* und die Zerlegung bewertet: es
    gewinnt das Zeichen, das den groessten zusammenhaengenden Block
    gleichbreiter Zeilen ergibt.  Prosa bildet keinen solchen Block, ein
    Briefkopf auch nicht -- eine Tabelle schon.
    """
    lines = [line for line in sample.splitlines()[:40] if line.strip()]
    if not lines:
        return ";"

    best, best_score = "", 0.0
    for candidate in _DELIMITERS:
        score = _table_score(lines, candidate)[0]
        if score > best_score:
            best, best_score = candidate, score
    if best:
        return best

    # Kein Kandidat bildet einen Block: es gibt hier keine Tabelle, die man
    # verderben koennte.  Der Sniffer darf noch einmal schauen, sonst bleibt
    # es beim in Deutschland ueblichen Semikolon.
    try:
        dialect = csv.Sniffer().sniff("\n".join(lines[:20]),
                                      delimiters="".join(_DELIMITERS))
        if dialect.delimiter in _DELIMITERS:
            return dialect.delimiter
    except csv.Error:
        logger.debug("CSV-Sniffer ohne Ergebnis, es bleibt beim Semikolon")
    return ";"
