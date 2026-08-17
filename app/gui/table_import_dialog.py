"""Auffang-Workflow 1: Tabelle einfuegen oder hochladen und selbst zuordnen.

Wann wird das gebraucht?
------------------------
Die automatische Erkennung ist gut, aber jeder Lieferant baut sein Angebot
anders auf.  Wenn sie einmal nicht greift, darf der Anwender nicht in einer
Sackgasse stehen.  Dieser Dialog ist der schnellste Ausweg:

    Excel oeffnen  ->  Bereich markieren  ->  Strg+C  ->  hier Strg+V
    ->  Spalten per Auswahlfeld zuordnen  ->  fertig

Zusaetzlich laesst sich eine Datei direkt laden (XLSX/CSV) -- etwa wenn die
Kopfzeile so ungewoehnlich ist, dass die Automatik sie nicht findet.

Der wichtigste Teil: Die Zuordnung wird auf Wunsch als Lieferantenprofil
gespeichert.  Beim naechsten Angebot desselben Lieferanten sitzt sie dann
automatisch -- aus einem Einzelfall wird gelerntes Wissen.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from decimal import Decimal

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ..models.enums import FieldOrigin, SourceKind
from ..models.offer_position import OfferPosition
from ..utils.parsing import (
    normalize_material_number,
    normalize_uom,
    normalize_whitespace,
    parse_date,
    parse_decimal,
    parse_int,
    similarity,
)
from .style import Colors

logger = logging.getLogger(__name__)

#: Zuordenbare Rollen.  Schluessel = Feldname in ``OfferPosition``.
#: Die Reihenfolge bestimmt die Reihenfolge im Auswahlfeld.
ROLES: tuple[tuple[str, str], tuple] = (
    ("", "– ignorieren –"),
    ("position_number", "Positionsnummer"),
    ("material_number", "Material (unsere SAP-Nummer)"),
    ("vendor_material_number", "Lieferantenartikelnummer"),
    ("description", "Bezeichnung"),
    ("quantity", "Menge"),
    ("uom", "Mengeneinheit"),
    ("price", "Preis"),
    ("price_unit", "Preiseinheit"),
    ("currency", "Waehrung"),
    ("min_order_qty", "Mindestbestellmenge"),
    ("lead_time_days", "Lieferzeit (Tage)"),
    ("valid_from", "Gueltig ab"),
    ("remarks", "Bemerkung"),
)

#: Rollen, die je Tabelle nur einmal vergeben werden duerfen
UNIQUE_ROLES = {key for key, _ in ROLES if key}

#: Stichworte fuer einen ersten Vorschlag, welche Spalte was ist.
#: Bewusst breit gehalten -- der Anwender korrigiert ohnehin.
ROLE_HINTS: dict[str, tuple[str, ...]] = {
    "position_number": ("pos", "position", "lfd", "nr", "item", "zeile", "no"),
    "material_number": ("material", "matnr", "sap", "ihre artikel", "ihre art",
                        "kundenartikel", "kundenmaterial", "kd-art", "ihre nummer",
                        "customer part", "customer material", "your part", "your ref",
                        "teilenummer", "ihre materialnummer"),
    "vendor_material_number": ("artikel", "artikelnr", "art.-nr", "sachnummer",
                               "hersteller", "unsere art", "unsere artikel", "typ",
                               "supplier part", "our part", "part no", "bestellnummer"),
    "description": ("bezeichnung", "beschreibung", "benennung", "text", "kurztext",
                    "description", "artikelbezeichnung", "warenbezeichnung"),
    "quantity": ("menge", "anzahl", "stück", "stueck", "qty", "quantity", "bedarf"),
    "uom": ("me", "einheit", "mengeneinheit", "meh", "uom", "unit", "verpackung"),
    "price": ("preis", "ep", "einzelpreis", "netto", "stückpreis", "stueckpreis",
              "unit price", "price", "betrag", "wert"),
    "price_unit": ("preiseinheit", "pe", "peh", "per", "price unit", "je"),
    "currency": ("währung", "waehrung", "wkz", "currency", "curr"),
    "min_order_qty": ("mindest", "mbm", "moq", "min order", "mindestabnahme"),
    "lead_time_days": ("lieferzeit", "wbz", "lead time", "lieferfrist", "wiederbeschaffung"),
    "valid_from": ("gültig", "gueltig", "valid", "ab dem", "preis ab"),
    "remarks": ("bemerkung", "hinweis", "anmerkung", "remark", "note", "kommentar"),
}


@dataclass
class TableImportResult:
    """Ergebnis der manuellen Zuordnung."""

    positions: list[OfferPosition] = field(default_factory=list)
    #: Spaltenueberschrift -> Feldname, fuer das Lieferantenprofil
    column_map: dict[str, str] = field(default_factory=dict)
    remember: bool = False
    source_name: str = "Eingefuegte Tabelle"

    @property
    def count(self) -> int:
        return len(self.positions)


class TableImportDialog(QDialog):
    """Tabelle einfuegen/laden und Spalten von Hand zuordnen."""

    def __init__(self, settings, vendor_name: str = "", parent=None) -> None:
        super().__init__(parent)
        self.settings = settings
        self.vendor_name = vendor_name
        self._grid: list[list[str]] = []
        self._role_boxes: list[QComboBox] = []
        self.result_data = TableImportResult()

        self.setWindowTitle("Tabelle einfuegen oder laden")
        self.resize(1080, 680)
        self._build()

    # ------------------------------------------------------------------
    # Aufbau
    # ------------------------------------------------------------------
    def _build(self) -> None:
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        heading = QLabel("Tabelle selbst zuordnen")
        heading.setObjectName("Heading")
        layout.addWidget(heading)

        hint = QLabel(
            "Markieren Sie den Positionsbereich in Excel, kopieren Sie ihn mit "
            "<b>Strg+C</b> und fuegen Sie ihn hier mit <b>Strg+V</b> ein – oder laden "
            "Sie eine Datei. Ordnen Sie danach oben in jeder Spalte zu, was darin steht. "
            "Nicht benoetigte Spalten bleiben auf „ignorieren“.")
        hint.setWordWrap(True)
        hint.setTextFormat(Qt.TextFormat.RichText)
        hint.setObjectName("SubHeading")
        layout.addWidget(hint)

        # -- Werkzeugleiste ---------------------------------------------
        tools = QHBoxLayout()
        paste_button = QPushButton("Aus Zwischenablage einfuegen (Strg+V)")
        paste_button.setObjectName("Primary")
        paste_button.clicked.connect(self.paste_from_clipboard)
        tools.addWidget(paste_button)

        file_button = QPushButton("Datei laden ...")
        file_button.clicked.connect(self.load_file)
        tools.addWidget(file_button)

        tools.addSpacing(20)
        tools.addWidget(QLabel("Kopfzeile:"))
        self.header_row = QSpinBox()
        self.header_row.setRange(0, 50)
        self.header_row.setSpecialValueText("keine")
        self.header_row.setToolTip("Welche Zeile enthaelt die Ueberschriften? "
                                   "0 = die Tabelle hat keine Kopfzeile.")
        self.header_row.valueChanged.connect(self._header_row_changed)
        tools.addWidget(self.header_row)

        suggest_button = QPushButton("Zuordnung vorschlagen")
        suggest_button.setToolTip("Aus den Ueberschriften einen Vorschlag ableiten")
        suggest_button.clicked.connect(self.suggest_roles)
        tools.addWidget(suggest_button)

        clear_button = QPushButton("Leeren")
        clear_button.clicked.connect(self.clear)
        tools.addWidget(clear_button)
        tools.addStretch(1)
        layout.addLayout(tools)

        # -- Zuordnungsleiste -------------------------------------------
        self.role_container = QWidget()
        self.role_layout = QHBoxLayout(self.role_container)
        self.role_layout.setContentsMargins(0, 0, 0, 0)
        self.role_layout.setSpacing(2)
        layout.addWidget(self.role_container)

        # -- Vorschau ----------------------------------------------------
        self.table = QTableWidget(0, 0)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setDefaultSectionSize(24)
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        layout.addWidget(self.table, 1)

        # -- Fuss --------------------------------------------------------
        self.status_label = QLabel("Noch keine Daten eingefuegt.")
        layout.addWidget(self.status_label)

        self.remember_box = QCheckBox(
            "Diese Zuordnung fuer zukuenftige Angebote dieses Lieferanten merken")
        self.remember_box.setChecked(bool(self.vendor_name))
        self.remember_box.setEnabled(bool(self.vendor_name))
        if not self.vendor_name:
            self.remember_box.setToolTip(
                "Dafuer muss zuerst ein Lieferant bekannt sein.")
        layout.addWidget(self.remember_box)

        buttons = QDialogButtonBox()
        self.ok_button = QPushButton("Positionen uebernehmen")
        self.ok_button.setObjectName("Primary")
        self.ok_button.setEnabled(False)
        cancel = QPushButton("Abbrechen")
        buttons.addButton(self.ok_button, QDialogButtonBox.ButtonRole.AcceptRole)
        buttons.addButton(cancel, QDialogButtonBox.ButtonRole.RejectRole)
        buttons.accepted.connect(self._accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    # ------------------------------------------------------------------
    # Daten einlesen
    # ------------------------------------------------------------------
    def keyPressEvent(self, event) -> None:  # noqa: N802
        if event.matches(Qt.Key.Key_V | Qt.KeyboardModifier.ControlModifier) or (
                event.key() == Qt.Key.Key_V
                and event.modifiers() & Qt.KeyboardModifier.ControlModifier):
            self.paste_from_clipboard()
            return
        super().keyPressEvent(event)

    def paste_from_clipboard(self) -> None:
        from PySide6.QtWidgets import QApplication

        text = QApplication.clipboard().text()
        if not text.strip():
            QMessageBox.information(
                self, "Zwischenablage leer",
                "In der Zwischenablage ist kein Text. Markieren Sie den Bereich in "
                "Excel und druecken Sie dort Strg+C.")
            return
        self.set_grid(self._parse_text(text))

    def load_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Tabelle laden", "",
            "Tabellen (*.xlsx *.xlsm *.csv *.txt *.tsv);;Alle Dateien (*.*)")
        if not path:
            return
        try:
            grid = self._read_file(path)
        except Exception as exc:  # noqa: BLE001 - Anwenderfehler nie eskalieren
            logger.warning("Tabelle konnte nicht geladen werden (%s): %s", path, exc)
            QMessageBox.warning(
                self, "Datei nicht lesbar",
                "Die Datei konnte nicht gelesen werden.\n\n"
                "Tipp: Speichern Sie sie als .xlsx oder .csv, oder kopieren Sie den "
                "Bereich direkt aus Excel und fuegen ihn hier ein.")
            return
        self.result_data.source_name = path
        self.set_grid(grid)

    @staticmethod
    def _parse_text(text: str) -> list[list[str]]:
        """Eingefuegten Text in ein Raster zerlegen.

        Excel und Outlook liefern beim Kopieren Tabulatoren; aus PDFs und Mails
        kommen oft Semikolon oder mehrere Leerzeichen.  Alle drei werden
        unterstuetzt -- entschieden wird nach dem Trenner, der die gleich-
        maessigste Spaltenzahl ergibt.
        """
        lines = [line for line in text.replace("\r\n", "\n").split("\n") if line.strip()]
        if not lines:
            return []

        kandidaten = []
        for delimiter in ("\t", ";", "|", ","):
            rows = [line.split(delimiter) for line in lines]
            counts = [len(r) for r in rows]
            if max(counts) < 2:
                continue
            # Gleichmaessigkeit bewerten: viele Zeilen mit gleicher Spaltenzahl
            haeufigste = max(set(counts), key=counts.count)
            score = counts.count(haeufigste) / len(counts) * haeufigste
            kandidaten.append((score, delimiter, rows))

        # Fallback: Aufteilung an zwei oder mehr Leerzeichen
        import re

        rows = [re.split(r"\s{2,}", line.strip()) for line in lines]
        counts = [len(r) for r in rows]
        if max(counts) >= 2:
            haeufigste = max(set(counts), key=counts.count)
            score = counts.count(haeufigste) / len(counts) * haeufigste
            kandidaten.append((score, "  ", rows))

        if not kandidaten:
            return [[line] for line in lines]

        kandidaten.sort(key=lambda item: item[0], reverse=True)
        rows = kandidaten[0][2]
        breite = max(len(r) for r in rows)
        return [[cell.strip() for cell in row] + [""] * (breite - len(row)) for row in rows]

    @staticmethod
    def _read_file(path: str) -> list[list[str]]:
        lower = path.lower()
        if lower.endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook

            workbook = load_workbook(path, data_only=True, read_only=True)
            sheet = workbook.active
            grid: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                grid.append(["" if value is None else str(value).strip() for value in row])
            workbook.close()
            return [row for row in grid if any(cell for cell in row)]

        raw = b""
        with open(path, "rb") as handle:
            raw = handle.read()
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")

        try:
            dialect = csv.Sniffer().sniff(text[:4000], delimiters=";,\t|")
            reader = csv.reader(io.StringIO(text), dialect)
        except csv.Error:
            reader = csv.reader(io.StringIO(text), delimiter=";")
        return [[cell.strip() for cell in row] for row in reader if any(c.strip() for c in row)]

    # ------------------------------------------------------------------
    # Anzeige
    # ------------------------------------------------------------------
    def set_grid(self, grid: list[list[str]]) -> None:
        self._grid = grid
        if not grid:
            self.clear()
            return

        breite = max(len(row) for row in grid)
        grid = [row + [""] * (breite - len(row)) for row in grid]
        self._grid = grid

        self.table.setRowCount(len(grid))
        self.table.setColumnCount(breite)
        self.table.setHorizontalHeaderLabels(
            [f"Spalte {i + 1}" for i in range(breite)])
        for r, row in enumerate(grid):
            for c, value in enumerate(row):
                self.table.setItem(r, c, QTableWidgetItem(value))

        self._build_role_boxes(breite)
        self.header_row.setMaximum(len(grid))
        self.header_row.setValue(1 if self._looks_like_header(grid[0]) else 0)
        self.suggest_roles()
        self._update_status()

    def _build_role_boxes(self, count: int) -> None:
        while self.role_layout.count():
            item = self.role_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        self._role_boxes = []

        for index in range(count):
            box = QComboBox()
            for key, label in ROLES:
                box.addItem(label, key)
            box.setMinimumWidth(150)
            box.currentIndexChanged.connect(self._roles_changed)
            self.role_layout.addWidget(box)
            self._role_boxes.append(box)
        self.role_layout.addStretch(1)
        self._sync_role_widths()

    def _sync_role_widths(self) -> None:
        """Auswahlfelder ueber den Spalten ausrichten."""
        for index, box in enumerate(self._role_boxes):
            width = self.table.columnWidth(index)
            box.setFixedWidth(max(110, width))

    @staticmethod
    def _looks_like_header(row: list[str]) -> bool:
        """Erste Zeile Ueberschrift oder schon Daten?"""
        gefuellt = [cell for cell in row if cell]
        if not gefuellt:
            return False
        zahlen = sum(1 for cell in gefuellt if parse_decimal(cell) is not None)
        return zahlen <= len(gefuellt) / 3

    def _header_row_changed(self) -> None:
        self.suggest_roles()
        self._update_status()

    def _headers(self) -> list[str]:
        index = self.header_row.value() - 1
        if index < 0 or index >= len(self._grid):
            return [f"Spalte {i + 1}" for i in range(len(self._role_boxes))]
        return [normalize_whitespace(cell) or f"Spalte {i + 1}"
                for i, cell in enumerate(self._grid[index])]

    # ------------------------------------------------------------------
    # Zuordnung
    # ------------------------------------------------------------------
    def suggest_roles(self) -> None:
        """Vorschlag aus Ueberschriften und Spalteninhalten ableiten."""
        if not self._role_boxes:
            return
        headers = self._headers()
        vergeben: set[str] = set()

        for index, box in enumerate(self._role_boxes):
            header = headers[index].lower() if index < len(headers) else ""
            beste_rolle, bester_wert = "", 0.0
            for role, stichworte in ROLE_HINTS.items():
                if role in vergeben:
                    continue
                for wort in stichworte:
                    if wort in header:
                        wert = 0.9 + len(wort) / 100.0
                    else:
                        wert = similarity(header, wort) * 0.8
                    if wert > bester_wert:
                        beste_rolle, bester_wert = role, wert
            if bester_wert < 0.55:
                beste_rolle = self._guess_by_content(index)
                if beste_rolle in vergeben:
                    beste_rolle = ""
            if beste_rolle:
                vergeben.add(beste_rolle)
            position = box.findData(beste_rolle)
            box.setCurrentIndex(max(0, position))
        self._roles_changed()

    def _guess_by_content(self, column: int) -> str:
        """Ohne brauchbare Ueberschrift: aus den Werten schliessen."""
        werte = [row[column] for row in self._data_rows() if column < len(row)]
        werte = [w for w in werte if w][:25]
        if not werte:
            return ""

        def anteil(pruefung) -> float:
            return sum(1 for w in werte if pruefung(w)) / len(werte)

        if anteil(lambda w: parse_date(w) is not None) >= 0.7:
            return "valid_from"
        if anteil(lambda w: len(w) <= 4 and w.isalpha()) >= 0.7:
            return "uom"
        if anteil(lambda w: parse_decimal(w) is not None) >= 0.7:
            return ""      # zu mehrdeutig: Menge oder Preis -- Anwender entscheidet
        if anteil(lambda w: len(w) > 12) >= 0.6:
            return "description"
        return ""

    def _roles_changed(self) -> None:
        """Doppelte Zuordnungen verhindern und farblich markieren."""
        gesehen: dict[str, int] = {}
        doppelt: set[int] = set()
        for index, box in enumerate(self._role_boxes):
            role = box.currentData()
            if not role:
                continue
            if role in gesehen:
                doppelt.add(index)
                doppelt.add(gesehen[role])
            else:
                gesehen[role] = index

        for index, box in enumerate(self._role_boxes):
            if index in doppelt:
                box.setStyleSheet(f"QComboBox {{ background: {Colors.RED_BG}; }}")
                box.setToolTip("Diese Rolle ist mehrfach vergeben.")
            elif box.currentData():
                box.setStyleSheet(f"QComboBox {{ background: {Colors.GREEN_BG}; }}")
                box.setToolTip("")
            else:
                box.setStyleSheet("")
                box.setToolTip("")

        self._update_status(doppelt)

    def _data_rows(self) -> list[list[str]]:
        start = self.header_row.value()
        return self._grid[start:] if start < len(self._grid) else []

    def _update_status(self, doppelt: set[int] | None = None) -> None:
        doppelt = doppelt or set()
        rollen = {box.currentData() for box in self._role_boxes if box.currentData()}
        zeilen = len(self._data_rows())

        if not self._grid:
            self.status_label.setText("Noch keine Daten eingefuegt.")
            self.ok_button.setEnabled(False)
            return

        probleme = []
        if doppelt:
            probleme.append("eine Rolle ist mehrfach vergeben")
        if "material_number" not in rollen and "vendor_material_number" not in rollen:
            probleme.append("weder Material noch Lieferantenartikelnummer zugeordnet")
        if "price" not in rollen:
            probleme.append("kein Preis zugeordnet")

        if probleme:
            self.status_label.setText(f"{zeilen} Datenzeile(n) – noch offen: "
                                      + ", ".join(probleme))
            self.status_label.setStyleSheet(f"color: {Colors.AMBER};")
        else:
            self.status_label.setText(
                f"{zeilen} Datenzeile(n) bereit zur Uebernahme "
                f"({len(rollen)} Spalten zugeordnet)")
            self.status_label.setStyleSheet(f"color: {Colors.GREEN};")
        self.ok_button.setEnabled(bool(zeilen) and not doppelt and bool(rollen))

    def clear(self) -> None:
        self._grid = []
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self._build_role_boxes(0)
        self._update_status()

    # ------------------------------------------------------------------
    # Uebernahme
    # ------------------------------------------------------------------
    def _accept(self) -> None:
        positions = self.build_positions()
        if not positions:
            QMessageBox.information(
                self, "Keine Positionen",
                "Aus den zugeordneten Spalten liess sich keine Position bilden.\n\n"
                "Pruefen Sie, ob die Kopfzeile richtig eingestellt ist und ob "
                "Material bzw. Preis zugeordnet sind.")
            return
        self.result_data.positions = positions
        self.result_data.remember = self.remember_box.isChecked()
        headers = self._headers()
        self.result_data.column_map = {
            headers[index]: box.currentData()
            for index, box in enumerate(self._role_boxes)
            if box.currentData() and index < len(headers)
        }
        self.accept()

    def build_positions(self) -> list[OfferPosition]:
        """Aus Raster und Zuordnung Positionen bauen.

        Werte werden genauso streng geparst wie bei der automatischen
        Erkennung: Was sich nicht eindeutig lesen laesst, bleibt leer statt
        geraten zu werden.  Die Herkunft ist ``MANUAL`` -- der Anwender hat
        die Zuordnung schliesslich selbst vorgegeben.
        """
        zuordnung = {index: box.currentData()
                     for index, box in enumerate(self._role_boxes)
                     if box.currentData()}
        if not zuordnung:
            return []

        positions: list[OfferPosition] = []
        for zeilennummer, row in enumerate(self._data_rows(), start=1):
            werte = {feld: (row[index].strip() if index < len(row) else "")
                     for index, feld in zuordnung.items()}
            if not any(werte.values()):
                continue
            if self._ist_summenzeile(werte):
                continue

            position = OfferPosition(
                source_kind=SourceKind.MANUAL,
                source_hint=f"manuell zugeordnete Tabelle, Zeile {zeilennummer}",
                raw_text=" | ".join(cell for cell in row if cell)[:300],
            )
            leer = True
            for feld, roh in werte.items():
                if not roh:
                    continue
                wert = self._konvertiere(feld, roh)
                if wert in (None, ""):
                    continue
                position.set_field(feld, wert, FieldOrigin.MANUAL)
                leer = False
            if leer:
                continue
            # Ohne Material und ohne Preis ist es keine brauchbare Position
            if not (position.material_number or position.vendor_material_number
                    or position.description):
                continue
            positions.append(position)

        logger.info("Manuelle Tabellenzuordnung: %d Position(en) aus %d Zeile(n)",
                    len(positions), len(self._data_rows()))
        return positions

    @staticmethod
    def _ist_summenzeile(werte: dict[str, str]) -> bool:
        text = " ".join(werte.values()).lower()
        marker = ("summe", "gesamt", "total", "zwischensumme", "subtotal", "mwst",
                  "vat", "netto gesamt", "zzgl", "endbetrag")
        return any(wort in text for wort in marker)

    @staticmethod
    def _konvertiere(feld: str, roh: str):
        if feld in ("quantity", "price", "min_order_qty"):
            return parse_decimal(roh)
        if feld in ("price_unit", "lead_time_days"):
            wert = parse_int(roh)
            return wert if wert and wert > 0 else None
        if feld == "valid_from":
            return parse_date(roh)
        if feld == "material_number":
            return normalize_material_number(roh)
        if feld == "uom":
            return normalize_uom(roh)
        if feld == "currency":
            return roh.upper()[:5]
        return normalize_whitespace(roh)
